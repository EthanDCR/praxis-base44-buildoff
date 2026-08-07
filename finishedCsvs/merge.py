#!/usr/bin/env python3
"""
Merge all data sources from Bulk Data Merge.xlsx into two normalized CSVs:
  - targets_import.csv   (one row per unique address)
  - contacts_import.csv  (one row per unique phone/email per address, scored by frequency)

Phone/email scoring: if the same value appears in N sources, score = N.
Contacts are sorted by score descending within each target (display_order).
"""

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found. Use the hail-pipeline venv.")
    sys.exit(1)

XLSX = Path(__file__).parent / "Bulk Data Merge.xlsx"
OUT_DIR = Path(__file__).parent

# ── Normalization ─────────────────────────────────────────────────────────────

STREET_ABBREVS = {
    r'\bstreet\b': 'st', r'\broad\b': 'rd', r'\bboulevard\b': 'blvd',
    r'\bavenue\b': 'ave', r'\bdrive\b': 'dr', r'\bcourt\b': 'ct',
    r'\blane\b': 'ln', r'\bplace\b': 'pl', r'\bcircle\b': 'cir',
    r'\btrail\b': 'trl', r'\bparkway\b': 'pkwy', r'\bway\b': 'wy',
    r'\bnorth\b': 'n', r'\bsouth\b': 's', r'\beast\b': 'e', r'\bwest\b': 'w',
    r'\bnortheast\b': 'ne', r'\bnorthwest\b': 'nw',
    r'\bsoutheast\b': 'se', r'\bsouthwest\b': 'sw',
}

def norm_street(s):
    s = str(s or '').lower().strip()
    for pat, rep in STREET_ABBREVS.items():
        s = re.sub(pat, rep, s)
    s = re.sub(r'[^a-z0-9\s]', '', s)
    return re.sub(r'\s+', ' ', s).strip()

def norm_zip(z):
    return str(z or '').strip().split('-')[0].zfill(5)

def addr_key(street, zip_code):
    return (norm_street(street), norm_zip(zip_code))

def norm_phone(p):
    digits = re.sub(r'\D', '', str(p or ''))
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    return digits if len(digits) == 10 else None

def norm_email(e):
    e = str(e or '').strip().lower()
    return e if '@' in e and '.' in e.split('@')[-1] else None

def str_val(v):
    if v is None:
        return ''
    return str(v).strip()

def num_val(v):
    try:
        f = float(str(v).replace(',', ''))
        return f if f == f else None  # NaN check
    except (ValueError, TypeError):
        return None

# ── Data structures ───────────────────────────────────────────────────────────

class TargetRecord:
    def __init__(self):
        self.street = ''
        self.city = ''
        self.state = ''
        self.zip = ''
        self.lat = None
        self.lng = None
        self.property_type = ''
        self.year_built = ''
        self.roof_type = ''
        self.roof_size_sq = None
        self.building_sqft = None
        self.stories = None
        self.hail_score_tier = ''
        self.hail_score_value = None
        self.largest_hail_inches = None
        self.largest_hail_date = ''
        self.most_recent_event_date = ''
        self.total_events_10yr = None
        self.data_sources = set()
        # phone/email: {normalized_value: {names: set, phone_type: str, tcpa: bool, sources: set}}
        self.phones = {}
        self.emails = {}

    def add_phone(self, raw, name='', phone_type='', tcpa=False, source=''):
        p = norm_phone(raw)
        if not p:
            return
        if p not in self.phones:
            self.phones[p] = {'names': set(), 'phone_type': phone_type, 'tcpa': tcpa, 'sources': set()}
        if name:
            self.phones[p]['names'].add(str_val(name))
        if phone_type:
            self.phones[p]['phone_type'] = phone_type
        if tcpa:
            self.phones[p]['tcpa'] = True
        if source:
            self.phones[p]['sources'].add(source)

    def add_email(self, raw, name='', source=''):
        e = norm_email(raw)
        if not e:
            return
        if e not in self.emails:
            self.emails[e] = {'names': set(), 'sources': set()}
        if name:
            self.emails[e]['names'].add(str_val(name))
        if source:
            self.emails[e]['sources'].add(source)

    def merge_property(self, **kwargs):
        for k, v in kwargs.items():
            if not v and v != 0:
                continue
            current = getattr(self, k, None)
            if not current and current != 0:
                setattr(self, k, v)

# ── Readers ───────────────────────────────────────────────────────────────────

def rows_from_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    return headers, rows[1:]

def read_deal_machine(ws, targets, source_label):
    headers, rows = rows_from_sheet(ws)
    h = {v: i for i, v in enumerate(headers)}

    def g(row, col):
        idx = h.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        street = str_val(g(row, 'property_address_street_1'))
        city   = str_val(g(row, 'property_address_city'))
        state  = str_val(g(row, 'property_address_state'))
        zip_   = norm_zip(g(row, 'property_address_zip_code'))
        if not street:
            continue

        key = addr_key(street, zip_)
        if key not in targets:
            targets[key] = TargetRecord()
        t = targets[key]
        t.data_sources.add(source_label)

        t.merge_property(
            street=street, city=city, state=state, zip=zip_,
            lat=num_val(g(row, 'latitude')),
            lng=num_val(g(row, 'longitude')),
            property_type=str_val(g(row, 'property_type')),
            roof_type=str_val(g(row, 'roof_cover') or g(row, 'roof_type')),
            building_sqft=num_val(g(row, 'living_area_sqft') or g(row, 'Building_Sq') or g(row, 'Building Sqft')),
            stories=num_val(g(row, 'stories')),
        )

        # Up to 5 contacts
        for n in range(1, 6):
            prefix = f'likely_owner_contact_{n}_'
            name = str_val(g(row, f'{prefix}full_name'))
            for p in range(1, 4):
                phone = g(row, f'{prefix}phone_number_{p}')
                ptype = str_val(g(row, f'{prefix}phone_number_{p}_phone_type'))
                t.add_phone(phone, name=name, phone_type=ptype, source=source_label)
            for e in range(1, 4):
                email = g(row, f'{prefix}email_address_{e}')
                t.add_email(email, name=name, source=source_label)

def read_master_export(ws, targets, source_label):
    headers, rows = rows_from_sheet(ws)
    h = {v: i for i, v in enumerate(headers)}

    def g(row, col):
        idx = h.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        street = str_val(g(row, 'Street'))
        city   = str_val(g(row, 'City'))
        state  = str_val(g(row, 'State'))
        zip_   = norm_zip(g(row, 'Zip'))
        if not street:
            continue

        key = addr_key(street, zip_)
        if key not in targets:
            targets[key] = TargetRecord()
        t = targets[key]
        t.data_sources.add(source_label)

        hail_date = str_val(g(row, 'Hail Date'))
        t.merge_property(
            street=street, city=city, state=state, zip=zip_,
            property_type=str_val(g(row, 'Property Type') or g(row, 'Building Type')),
            year_built=str_val(g(row, 'Year Built')),
            roof_type=str_val(g(row, 'Roof Types')),
            roof_size_sq=num_val(g(row, 'Roof Size (Sq)')),
            building_sqft=num_val(g(row, 'Building Area (sqft)')),
            largest_hail_inches=num_val(g(row, 'Hail Size (in)')),
            largest_hail_date=hail_date,
            most_recent_event_date=hail_date,
        )

        # Up to 20 owners (O1-O20)
        for n in range(1, 21):
            prefix = f'O{n}_'
            name = str_val(g(row, f'{prefix}Name'))
            if not name:
                continue
            personal_phone = g(row, f'{prefix}Personal_Phone')
            t.add_phone(personal_phone, name=name, source=source_label)
            personal_email = g(row, f'{prefix}Personal_Email')
            t.add_email(personal_email, name=name, source=source_label)
            # Numbered phones P1-P12
            for p in range(1, 13):
                phone = g(row, f'{prefix}P{p}_Number')
                ptype = str_val(g(row, f'{prefix}P{p}_Type'))
                t.add_phone(phone, name=name, phone_type=ptype, source=source_label)
            # Numbered emails E1-E17
            for e in range(1, 18):
                email = g(row, f'{prefix}E{e}_Email')
                t.add_email(email, name=name, source=source_label)

def read_hail_results(ws, targets, source_label):
    headers, rows = rows_from_sheet(ws)
    h = {v: i for i, v in enumerate(headers)}

    def g(row, col):
        idx = h.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        street = str_val(g(row, 'street'))
        zip_   = norm_zip(g(row, 'zip'))
        if not street:
            continue

        key = addr_key(street, zip_)
        if key not in targets:
            targets[key] = TargetRecord()
        t = targets[key]
        t.data_sources.add(source_label)

        t.merge_property(
            street=street,
            city=str_val(g(row, 'city')),
            state=str_val(g(row, 'state')),
            zip=zip_,
            lat=num_val(g(row, 'lat')),
            lng=num_val(g(row, 'lon')),
            hail_score_tier=str_val(g(row, 'score_tier')),
            hail_score_value=num_val(g(row, 'score_value')),
            largest_hail_inches=num_val(g(row, 'largest_hail_inches')),
            largest_hail_date=str_val(g(row, 'largest_hail_date')),
            most_recent_event_date=str_val(g(row, 'most_recent_event_date')),
            total_events_10yr=num_val(g(row, 'total_events_10yr')),
        )

def read_skiptrace(ws, targets, source_label):
    headers, rows = rows_from_sheet(ws)
    h = {v: i for i, v in enumerate(headers)}

    def g(row, col):
        idx = h.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        street = str_val(g(row, 'street'))
        zip_   = norm_zip(g(row, 'zip'))
        if not street:
            continue

        key = addr_key(street, zip_)
        if key not in targets:
            targets[key] = TargetRecord()
        t = targets[key]
        t.data_sources.add(source_label)

        t.merge_property(
            street=street,
            city=str_val(g(row, 'city')),
            state=str_val(g(row, 'state')),
            zip=zip_,
            lat=num_val(g(row, 'lat')),
            lng=num_val(g(row, 'lon')),
            hail_score_tier=str_val(g(row, 'score_tier')),
            hail_score_value=num_val(g(row, 'score_value')),
            largest_hail_inches=num_val(g(row, 'largest_hail_inches')),
            largest_hail_date=str_val(g(row, 'largest_hail_date')),
            most_recent_event_date=str_val(g(row, 'most_recent_event_date')),
            total_events_10yr=num_val(g(row, 'total_events_10yr')),
        )

        # Up to 3 persons
        for n in range(1, 4):
            prefix = f'p{n}_'
            name = str_val(g(row, f'{prefix}name'))
            for p in range(1, 6):
                phone = g(row, f'{prefix}phone{p}')
                ptype = str_val(g(row, f'{prefix}phone{p}_type'))
                tcpa_raw = g(row, f'{prefix}phone{p}_tcpa')
                tcpa = str(tcpa_raw).lower() in ('true', '1', 'yes') if tcpa_raw else False
                t.add_phone(phone, name=name, phone_type=ptype, tcpa=tcpa, source=source_label)
            for e in range(1, 4):
                email = g(row, f'{prefix}email{e}')
                t.add_email(email, name=name, source=source_label)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading {XLSX.name}...")
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    targets = {}  # addr_key -> TargetRecord

    # Sheet 1 & 2 are the same DealMachine export — only read once
    print("Reading Deal Machine...")
    read_deal_machine(wb['Deal Machine Contact Export #1'], targets, 'DealMachine')

    print("Reading Master Export (SalesMastery)...")
    read_master_export(wb['Master Export - TX OK KS MO NM '], targets, 'SalesMastery')

    print("Reading Hail Results...")
    read_hail_results(wb['adressFatList_hail_results'], targets, 'HailPipeline')

    print("Reading Skip Trace...")
    read_skiptrace(wb['skiptrace_8497_results'], targets, 'BatchData')

    print(f"\nTotal unique addresses: {len(targets)}")

    # ── Write targets_import.csv ──────────────────────────────────────────────
    target_fields = [
        'import_key', 'street', 'city', 'state', 'zip',
        'lat', 'lng', 'property_type', 'year_built',
        'roof_type', 'roof_size_sq', 'building_sqft', 'stories',
        'hail_score_tier', 'hail_score_value', 'largest_hail_inches',
        'largest_hail_date', 'most_recent_event_date', 'total_events_10yr',
        'data_sources',
    ]

    targets_path = OUT_DIR / 'targets_import.csv'
    with open(targets_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=target_fields)
        w.writeheader()
        for i, (key, t) in enumerate(targets.items()):
            w.writerow({
                'import_key':            f"{key[0]}|{key[1]}",
                'street':                t.street,
                'city':                  t.city,
                'state':                 t.state,
                'zip':                   t.zip,
                'lat':                   t.lat if t.lat is not None else '',
                'lng':                   t.lng if t.lng is not None else '',
                'property_type':         t.property_type,
                'year_built':            t.year_built,
                'roof_type':             t.roof_type,
                'roof_size_sq':          t.roof_size_sq if t.roof_size_sq is not None else '',
                'building_sqft':         t.building_sqft if t.building_sqft is not None else '',
                'stories':               t.stories if t.stories is not None else '',
                'hail_score_tier':       t.hail_score_tier,
                'hail_score_value':      t.hail_score_value if t.hail_score_value is not None else '',
                'largest_hail_inches':   t.largest_hail_inches if t.largest_hail_inches is not None else '',
                'largest_hail_date':     t.largest_hail_date,
                'most_recent_event_date': t.most_recent_event_date,
                'total_events_10yr':     t.total_events_10yr if t.total_events_10yr is not None else '',
                'data_sources':          '|'.join(sorted(t.data_sources)),
            })

    print(f"Written: {targets_path} ({len(targets)} rows)")

    # ── Write contacts_import.csv ─────────────────────────────────────────────
    contact_fields = [
        'import_key', 'contact_name', 'value', 'contact_type',
        'phone_type', 'tcpa', 'score', 'sources', 'display_order',
    ]

    contacts_path = OUT_DIR / 'contacts_import.csv'
    total_contacts = 0

    with open(contacts_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=contact_fields)
        w.writeheader()

        for key, t in targets.items():
            import_key = f"{key[0]}|{key[1]}"
            contacts = []

            for phone, meta in t.phones.items():
                contacts.append({
                    'import_key':    import_key,
                    'contact_name':  ', '.join(sorted(meta['names'])),
                    'value':         phone,
                    'contact_type':  'phone',
                    'phone_type':    meta['phone_type'],
                    'tcpa':          str(meta['tcpa']).lower(),
                    'score':         len(meta['sources']),
                    'sources':       '|'.join(sorted(meta['sources'])),
                    'display_order': 0,
                })

            for email, meta in t.emails.items():
                contacts.append({
                    'import_key':    import_key,
                    'contact_name':  ', '.join(sorted(meta['names'])),
                    'value':         email,
                    'contact_type':  'email',
                    'phone_type':    '',
                    'tcpa':          '',
                    'score':         len(meta['sources']),
                    'sources':       '|'.join(sorted(meta['sources'])),
                    'display_order': 0,
                })

            # Sort by score desc, then type (phones first), then value
            contacts.sort(key=lambda c: (-c['score'], c['contact_type'], c['value']))
            for i, c in enumerate(contacts):
                c['display_order'] = i + 1
                w.writerow(c)
                total_contacts += 1

    print(f"Written: {contacts_path} ({total_contacts} rows)")
    print(f"\nDone. {len(targets)} targets, {total_contacts} contacts.")
    print(f"Average contacts per target: {total_contacts / len(targets):.1f}")

if __name__ == '__main__':
    main()
