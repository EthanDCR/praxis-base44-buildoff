import json, os, sys, time
import requests
import openpyxl
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed

load_dotenv('hail-pipeline/.env')

APP_ID  = os.getenv('BASE44_APP_ID')
API_KEY = os.getenv('BASE44_API_KEY')
BASE    = f'https://base44.app/api/apps/{APP_ID}/entities'
HEADERS = {'Content-Type': 'application/json', 'api_key': API_KEY}

FILES = ['testCsvs/hailExport1.xlsx', 'testCsvs/hailExport2.xlsx']
LIST_NAME = '7-28-26 IL Storm'

def parse_file(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows

def split_field(val):
    if not val:
        return []
    return [v.strip() for v in str(val).split(';') if v.strip()]

def build_target(row, list_id):
    city  = row.get('City') or ''
    state = row.get('State') or ''
    zip_  = row.get('ZIP') or ''
    line2 = ', '.join(filter(None, [city, state])) + (f' {zip_}' if zip_ else '')

    phones = split_field(row.get('Owner Phone(s)'))
    emails = split_field(row.get('Owner Email(s)'))
    owner  = row.get('Owner Name') or ''

    contact = {'name': owner, 'phones': phones, 'emails': emails, 'company': '', 'title': ''} if (owner or phones) else None

    return {
        'list_id':       list_id,
        'line1':         row.get('Address') or '',
        'line2':         line2.strip(', '),
        'lat':           row.get('Latitude'),
        'lng':           row.get('Longitude'),
        'property_type': row.get('Property Type'),
        'year_built':    str(row.get('Year Built')) if row.get('Year Built') else None,
        'status':        'new',
        'contacts':      [contact] if contact else [],
        'is_sample':     False,
    }

def create_target(payload):
    r = requests.post(f'{BASE}/Target', headers=HEADERS, json=payload, timeout=20)
    r.raise_for_status()
    return r.json()

def main():
    # Parse both files
    all_rows = []
    for f in FILES:
        rows = parse_file(f)
        print(f'{f}: {len(rows)} rows')
        all_rows.extend(rows)
    print(f'Total: {len(all_rows)} rows')

    # Create the list
    r = requests.post(f'{BASE}/CallList', headers=HEADERS, json={'name': LIST_NAME, 'status': 'in_progress'})
    r.raise_for_status()
    list_id = r.json()['id']
    print(f'Created list "{LIST_NAME}" (id={list_id})')

    # Build payloads
    payloads = [build_target(row, list_id) for row in all_rows]

    # Push sequentially with delay to respect rate limits
    success, failure = 0, 0
    start = time.time()

    for i, p in enumerate(payloads):
        try:
            create_target(p)
            success += 1
        except Exception as e:
            failure += 1
            print(f'  Failed row {i}: {e}')
        time.sleep(0.35)
        if (i + 1) % 100 == 0:
            elapsed = time.time() - start
            print(f'  {i+1}/{len(payloads)} done ({elapsed:.0f}s, ~{(len(payloads)-(i+1))*0.35/60:.1f}min left)')

    print(f'\nDone — {success} created, {failure} failed ({time.time()-start:.0f}s)')

if __name__ == '__main__':
    main()
